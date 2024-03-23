import os
from PyPDF2 import PdfReader
from openpyxl import Workbook
import re
from words2numsrus.extractor import NumberExtractor
import traceback
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
extractor = NumberExtractor()
# Create an Excel workbook
workbook = Workbook()
sheet = workbook.active

# Directory containing your invoice files
directory = 'C:/Users/acer/OneDrive/Desktop/Dusel-Projects'

# Iterate over each file in the directory
for filename in os.listdir(directory):
    if filename.endswith('.pdf'):
        file_path = os.path.join(directory, filename)

        # Open the PDF file
        with open(file_path, 'rb') as file:
            try:
                pdf_reader = PdfReader(file)

                # Extract text from the first page
                first_page = pdf_reader.pages[0]
                text = first_page.extract_text()
                if len(pdf_reader.pages) > 1:
                    text += pdf_reader.pages[1].extract_text()

                # print(text)
                # print('/'*88)

                commission_agent_match = re.search(r'Поставщик:\s*(.*)', text)
                commission_agent = commission_agent_match.group(
                    1) if commission_agent_match else None

                if commission_agent is None:
                    commission_agent = re.search(r'Комиссионер:\s*(.*)', text)
                    commission_agent = commission_agent.group(
                        1) if buyer_match else None

                # Find the "Покупатель" data
                buyer_match = re.search(r'Покупатель:\s*(.*)', text)
                buyer = buyer_match.group(1) if buyer_match else None

                service = re.search(
                    r'Оказание услуг\n([\s\S]+?)\nуслуга \(сум\)', text)
                text_service = service.group(1).split(
                    "\n")[-1] if service else None

                if service is None:
                    service = re.search(
                        r'.*( – Услуги| – Услуга|Транспортно-\nэкспедиторская услуга).*', text)
                    text_service = service.group(0) if service else None

                if service is None:
                    service = re.search(
                        r'.* – .*', text)
                    text_service = service.group(0) if service else None

                # Find the "Итого" data
                total_match = re.search(r'Всего к оплате:\s*(.*)', text)
                total = total_match.group(1) if total_match else None
                total = extractor.replace_groups(total).split(' ')[0]

                match = re.search(r'ОТПРАВЛЕНО[^\n]*\n([^ПОДТВЕРЖДЁН]*)', text)
                text_after_operator = match.group(1) if match else None

                # Extract the date from the text after оператор
                date_match = re.search(
                    r'\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}:\d{2}', text_after_operator)

                date = date_match.group(0).split(
                    ' ')[0] if date_match else None

                if buyer is not None and commission_agent is not None and total is not None and date_match is not None and text_service is not None:

                    clean_buyer = re.findall(r'"([^"]+)"', buyer)[0]
                    sheet = workbook[clean_buyer] if clean_buyer in workbook.sheetnames else workbook.create_sheet(
                        clean_buyer)

                    # Define column names in the sheet
                    column_names = ['Поставщик', 'услуг', 'Итого', 'date']

                    # If the sheet is newly created, append the column names
                    if sheet.max_row == 1:
                        sheet.append(column_names)

                    sheet.append([commission_agent, text_service, total, date])
                    file.close()
                    os.remove(file_path)

                else:
                    file.close()

                print("Поставщик:", commission_agent)
                print("Покупатель:", buyer)
                print("услуг:", text_service)
                print("Итого:", total)
                print("Date:", date)
                print('-'*88)
            except Exception as e:
                print(e)
                traceback.print_exc()


sheet = workbook["Sheet"]

workbook.remove(sheet)

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    # data_range = sheet[2:sheet.max_row]
    # sorted_rows = sorted(data_range, key=lambda x: x[0].value)
    # sheet.delete_rows(2, sheet.max_row)

    # for row in sorted_rows:
    #     sheet.append([cell.value for cell in row])


workbook.save('C:/Users/acer/OneDrive/Desktop/Dusel-Projects/file.xlsx')
