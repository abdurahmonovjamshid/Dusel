from openpyxl import load_workbook

# Load the workbook
workbook = load_workbook('file.xlsx')

# Select the desired sheet
sheet = workbook.active

# Get the range of cells containing data
data_range = sheet[1:sheet.max_row]

# Sort the rows based on column A values
sorted_rows = sorted(data_range, key=lambda x: x[0].value)

# Clear the existing content in the sheet
sheet.delete_rows(1, sheet.max_row)

# Write the sorted rows back to the sheet
for row in sorted_rows:
    sheet.append([cell.value for cell in row])

# Save the workbook with sorted data
workbook.save('sorted_file.xlsx')